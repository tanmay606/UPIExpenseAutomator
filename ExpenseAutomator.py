import os
import datetime
import imaplib
import email
import re
import pandas as pd
from email.header import decode_header
from openpyxl import load_workbook


LOG_FILE = os.path.join(os.path.dirname(__file__), "log.txt")  # Log file path

# Gmail credentials
EMAIL_USER = ""  # Replace with your Gmail
EMAIL_PASS = ""# Use App Password if 2FA is enabled

# Path to Excel file
EXCEL_FILE = os.path.join(os.path.dirname(__file__), "Expenses", "Feb 25.xlsx")
SHEET_NAME = "Feb 25"  # Target sheet for February transactions

# Get current date
TODAY_DATE = int(datetime.datetime.today().strftime("%d"))  # Format: Day (DD)
GMAIL_SEARCH_DATE = datetime.datetime.today().strftime("%d-%b-%Y")  # Format: DD-MMM-YYYY ( configured to fetch transactions of today's date, can make changes to fetch from certain date till today.)

# Expense categories (Added "Other Expenses" as 12th category)
EXPENSE_CATEGORIES = {
    1: "Food",
    2: "Travel",
    3: "Rent & Electricity",
    4: "Grooming Expense",
    5: "EMI Expense",
    6: "Indore Expense",
    7: "Subscription Based Expense",
    8: "Clothing Expense",
    9: "Business Related Expense",
    10: "Donation Expense",
    11: "Personal Expense",
    12: "Other Expenses"
}


# Connect to Gmail
def connect_gmail():
    try:
        mail = imaplib.IMAP4_SSL("imap.gmail.com")
        mail.login(EMAIL_USER, EMAIL_PASS)
        mail.select("inbox")
        return mail
    except Exception as e:
        print(f"❌ Error connecting to Gmail: {e}")
        exit()


# Log file path
LOG_FILE = os.path.join(os.path.dirname(__file__), "log.txt")

# Fetch UPI transaction emails
def fetch_upi_emails(mail):
    search_query = f'(FROM "alerts@hdfcbank.net" SINCE "{GMAIL_SEARCH_DATE}")'
    result, data = mail.search(None, search_query)
    email_ids = data[0].split()
    transactions = []

    print(f"\n📩 Total emails from HDFC today ({GMAIL_SEARCH_DATE}): {len(email_ids)}\n")

    # Regex pattern for UPI transactions
    upi_pattern = r"Rs\.\s?(\d+\.\d{2})\s?has been debited .*? to VPA (\S+)\s(.+?) on (\d{2}-\d{2}-\d{2})"

    for e_id in email_ids:
        try:
            result, msg_data = mail.fetch(e_id, "(RFC822)")
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            # Decode subject
            subject, encoding = decode_header(msg["Subject"])[0]
            if isinstance(subject, bytes):
                subject = subject.decode(encoding or "utf-8", errors="ignore")

            email_text = ""

            # Extract email body
            if msg.is_multipart():
                for part in msg.walk():
                    content_type = part.get_content_type()
                    content_disposition = str(part.get("Content-Disposition"))

                    if content_type == "text/plain" and "attachment" not in content_disposition:
                        email_text = part.get_payload(decode=True).decode("utf-8", errors="ignore")
                        break  # Prefer text/plain
                    elif content_type == "text/html" and not email_text:
                        email_text = part.get_payload(decode=True).decode("utf-8", errors="ignore")
            else:
                email_text = msg.get_payload(decode=True).decode("utf-8", errors="ignore")

            # Extract UPI transaction details
            matches = re.findall(upi_pattern, email_text)

            if matches:
                for match in matches:
                    amount = float(match[0])
                    vpa_id = match[1]
                    party_name = match[2].strip().lower()
                    date = match[3]

                    transactions.append({"date": date, "amount": amount, "vpa_id": vpa_id, "party_name": party_name})
                    print(f"✅ Amount: Rs.{amount}, UPI ID: {vpa_id}, Party: {party_name}, Date: {date}")
            else:
                pass
                #print("❌ No UPI transaction details found in this email.\n")

        except Exception as e:
            print(f"❌ Error processing email: {e}")

    return transactions  

from datetime import datetime

def update_excel(transactions):
    """
    Categorizes transactions, updates the 'Feb 25' and 'Daily 2025' sheets,
    prevents duplicates, logs balance changes, updates O1 with current date-time, 
    and provides a category-wise summary of total amounts processed.
    """

    total_upi_amount = round(sum(txn['amount'] for txn in transactions), 2)  # Total from UPI transactions
    total_amount_added = 0  # Total added to categories
    total_amount_skipped = 0  # Total skipped transactions
    category_sums = {}  # Stores total added for each category
    log_entries = []  
    timestamp = datetime.now().strftime("%d-%b-%Y %H:%M:%S")  # Log timestamp

    # Generate formatted date and time for O1
    formatted_datetime = datetime.now().strftime("%d-%b-%Y %I:%M%p")

    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb[SHEET_NAME]
        ws_daily = wb["Daily 2025"]  # Restore Daily Expense Sheet
    except FileNotFoundError:
        print("❌ Excel file not found. Please create the file first.")
        return

    # **Step 1: Read previous total amount from log file (last run total expense)**
    last_run_total = 0.0
    try:
        with open(LOG_FILE, "r") as log_file:
            lines = log_file.readlines()
            for line in reversed(lines):
                if "Total Expense Added Today" in line:
                    last_run_total = round(float(line.split(":")[1].strip().replace("Rs.", "").strip()), 2)
                    break
    except (FileNotFoundError, ValueError):
        pass  # If log file doesn't exist or is corrupted, ignore and proceed

    # **Step 2: Compare UPI Total with Last Run Total to Prevent Duplicates**
    print(f"📊 Mail Expense Total: Rs.{total_upi_amount:.2f}, Log File Total: Rs.{last_run_total:.2f}")
    if total_upi_amount == last_run_total:
        print(f"⚠️ Expenses for today have already been added. Skipping this run.")
        with open(LOG_FILE, "a") as log_file:
            log_file.write(f"\n[{timestamp}] ⚠️ SKIPPED: Expenses already recorded. No duplicate booking.\n")
            log_file.write("\n" + "=" * 50 + "\n\n")  # Add separator after logs
        return  # **Skip processing and exit**

    # **Update cell O1 with the current date and time**
    ws["O1"].value = formatted_datetime

    print(f"📌 Please categorize each transaction:")

    for txn in transactions:
        print(f"\n💰 Amount: Rs.{txn['amount']:.2f}")
        print(f"👤 Party: {txn['party_name']}")

        for key, value in EXPENSE_CATEGORIES.items():
            print(f"{key}: {value}")

        while True:
            user_input = input("Enter category number (Press ENTER to SKIP): ").strip()
            if user_input == "":
                print("⚠️ Skipping this transaction...\n")
                log_entries.append(f"[{timestamp}] Transaction SKIPPED: Rs.{txn['amount']:.2f} for '{txn['party_name']}'")
                total_amount_skipped += txn["amount"]
                break
            try:
                category_choice = int(user_input)
                if category_choice in EXPENSE_CATEGORIES:
                    category_name = EXPENSE_CATEGORIES[category_choice]
                    break
                else:
                    print("⚠️ Invalid input! Please enter a valid category number.")
            except ValueError:
                print("⚠️ Please enter a number.")

        if user_input == "":
            continue  # Skip this transaction and move to the next

        # Find the row corresponding to the category
        category_row = None
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == category_name:
                category_row = row
                break

        if category_row is None:
            print(f"❌ Category '{category_name}' not found in the Excel sheet.")
            continue

        # **Step 3: Update Food Expenses in "Daily 2025"**
        if category_name == "Food":
            prev_food_expense = 0.0  # Default if no previous value
            date = txn["date"]
            day = int(date.split("-")[0])  # Extract the day (DD)
            month_name = datetime.strptime(date, "%d-%m-%y").strftime("%B")  # Convert to full month name

            # Find the column corresponding to the month
            month_col = next((col for col in range(2, ws_daily.max_column + 1)
                              if ws_daily.cell(row=2, column=col).value == month_name), None)
            if month_col is None:
                continue  # Skip if month column is not found

            # Find the row corresponding to the day
            day_row = next((row for row in range(3, ws_daily.max_row + 1)
                            if isinstance(ws_daily.cell(row=row, column=1).value, (int, float)) and
                            ws_daily.cell(row=row, column=1).value == day), None)
            if day_row is None:
                continue  # Skip if day row is not found

            # Get previous food expense, default to 0 if empty
            prev_food_expense = float(ws_daily.cell(row=day_row, column=month_col).value or 0)

            # Update the cell with the new food expense
            new_food_expense = prev_food_expense + txn["amount"]
            ws_daily.cell(row=day_row, column=month_col, value=new_food_expense)

            # ✅ **LOG THE FOOD EXPENSE UPDATE**
            total_amount_added += txn["amount"]
            category_sums["Food"] = category_sums.get("Food", 0) + txn["amount"]

            log_entries.append(
                f"[{timestamp}] 'Daily 2025' (Food) -> Previous: Rs.{prev_food_expense:.2f}, Added: Rs.{txn['amount']:.2f}, New: Rs.{new_food_expense:.2f}"
            )

        # **Step 4: Update Non-Food Expenses (Preserve Formulas)**
        else:
            cell_value = ws.cell(row=category_row, column=3).value
            prev_balance = 0.0  # Initialize prev_balance with a default value

            # **Check if the cell contains a formula and preserve it**
            if isinstance(cell_value, str) and cell_value.startswith("="):
                updated_formula = f"{cell_value} + {txn['amount']}"
                ws.cell(row=category_row, column=3, value=updated_formula)
                try:
                    prev_balance = float(eval(cell_value[1:]))  # Extract numerical part safely
                except:
                    prev_balance = 0.0  # Fallback in case of error in evaluation
                new_balance = prev_balance + txn["amount"]
            else:
                prev_balance = float(cell_value) if isinstance(cell_value, (int, float)) else 0.0
                new_balance = prev_balance + txn["amount"]
                ws.cell(row=category_row, column=3, value=new_balance)

            # ✅ **LOG THE NON-FOOD EXPENSE UPDATE**
            total_amount_added += txn["amount"]
            category_sums[category_name] = category_sums.get(category_name, 0) + txn["amount"]

            log_entries.append(
                f"[{timestamp}] '{category_name}' -> Previous: Rs.{prev_balance:.2f}, Added: Rs.{txn['amount']:.2f}, New: Rs.{new_balance:.2f}"
            )




    # **Save and Log Updates**
    wb.save(EXCEL_FILE)
    wb.close()

    # Print Summary in Terminal
    print("\n📊 **Summary of Today's Transactions:**")
    print(f"💰 Total UPI Transactions from Mail: Rs. {total_upi_amount:.2f}")
    print(f"✅ Total Amount Added: Rs. {total_amount_added:.2f}")
    print(f"⚠️ Total Amount Skipped: Rs. {total_amount_skipped:.2f}")
    print("\n📂 **Category-wise Breakdown:**")
    for category, amount in category_sums.items():
        print(f"   - {category}: Rs. {amount:.2f}")

    with open(LOG_FILE, "a") as log_file:
        log_file.write("\n" + "=" * 50 + "\n")  # Separator for logs
        for entry in log_entries:
            log_file.write(entry + "\n")
        log_file.write(f"\nTotal Expense Added Today: Rs. {total_amount_added:.2f}\n")
        log_file.write(f"Total Amount Skipped: Rs. {total_amount_skipped:.2f}\n")
        log_file.write("\n📂 **Category-wise Breakdown:**\n")
        for category, amount in category_sums.items():
            log_file.write(f"   - {category}: Rs. {amount:.2f}\n")
        log_file.write("\n" + "=" * 50 + "\n\n")

    print(f"\n📂 Excel file updated successfully. Logs written to {LOG_FILE}")


    # Main execution
if __name__ == "__main__":
    print(f"🔍 Connecting to Gmail... Fetching transactions from/for {GMAIL_SEARCH_DATE}")
    mail = connect_gmail()

    print(f"📩 Fetching today's UPI transactions from HDFC emails...")
    transactions = fetch_upi_emails(mail)  

    mail.logout()  # Logout AFTER fetching emails

    if transactions:
        update_excel(transactions)
    else:
        print("❌ No UPI transactions to process.")